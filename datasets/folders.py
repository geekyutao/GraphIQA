import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook
from torchvision.transforms import Compose, ToTensor, CenterCrop, RandomCrop, Normalize, Resize
import pandas as pd
import lmdb
import six
import random

class LIVEFolder(data.Dataset):

    def __init__(self, root, index, patch_num, save,type_sel='all'):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                if type_sel == 'all':
                    for aug in range(patch_num):
                        sample.append((imgpath[item], labels[0][item]))
                else:
                    if type_sel in imgpath[item]:
                        for aug in range(patch_num):
                            sample.append((imgpath[item], labels[0][item]))
                # print(self.imgpath[item])
        self.samples = sample
        self.transform = self.img_transform(224)

        if save:
            header = ['image','dmos']
            csv_test = open(os.path.join(root, 'test-split.csv'), 'w')
            writer = csv.writer(csv_test)
            writer.writerow(header)

            # test_writer = csv.writer(csv_test)
            # test_writer.writerow(header)
            for i, line in enumerate(self.samples):
                writer.writerow(line)
            csv_test.close()

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])


class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index,  patch_num):

        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'images', imgpath[item][0][0]), labels[item]))

        self.samples = sample
        self.transform = self.img_transform(224)


    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
        ])

class KadidFolder(data.Dataset):
    def __init__(self,root, index, patch_num, sel):
        # imgname = []
        # mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        df = pd.read_csv(csv_file)
        imgnames = df['dist_img'].tolist()
        # print(imgnames)
        labels = np.array(df['dmos']).astype(np.float32)
        # print(labels)
        refname = np.unique(np.array(df['ref_img']))
        refnames_all = np.array(df['ref_img'])
        # print(refnames_all)
        sample = []
        for i, item in enumerate(index):
            # print(refname[index[i]])
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            # print(train_sel)
            for j, item in enumerate(train_sel):
                if sel =='all':
                    for aug in range(patch_num):
                        sample.append((os.path.join(root, 'images', imgnames[item]), labels[item]))
                elif sel == int(imgnames[item].split('_')[1]):
                    for aug in range(patch_num):
                        sample.append((os.path.join(root, 'images', imgnames[item]), labels[item]))

        self.samples = sample
        self.transform = self.img_transform(224)

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])

class KadisLMDBFolder(data.Dataset):
    def __init__(self, db_path, shuffle=True):
        self.db_path = db_path + '/dist_imgs_lmdb_237'
        self.env = lmdb.open(self.db_path,
                             max_readers=1,
                             readonly=True,
                             lock=False,
                             readahead=False,
                             meminit=False)
        root = db_path
        csv_file = os.path.join(root, 'dmos.csv')
        df = pd.read_csv(csv_file)
        imgnames = df['dist_img'].tolist()
        labels = np.array(df['type']).astype(np.float32)
        sample = []
        for i, item in enumerate(imgnames):
            sample.append((os.path.join(root, 'dist_imgs', item), labels[i]))

        self.txn = self.env.begin(write=False)
        # self.nSamples = int(self.txn.get('num-samples'.encode()))
        self.nSamples = len(sample)
        # self.indices = range(self.nSamples)
        self.samples = sample
        if shuffle:
            random.shuffle(self.samples)

        self.transform = self.img_transform(224)

    def __getitem__(self, index):
        path, target = self.samples[index]
        imgKey = path
        imageBin = self.txn.get(imgKey.encode())
        buf = six.BytesIO()
        buf.write(imageBin)
        buf.seek(0)
        img = Image.open(buf).convert('RGB')
        img = self.transform(img)

        return img, target

    def __len__(self):
        return self.nSamples

    def __repr__(self):
        return self.__class__.__name__ + ' (' + self.db_path + ')'

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])

class KadisFolder(data.Dataset):
    def __init__(self,root, index, patch_num, sel):
        # imgname = []
        # mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        df = pd.read_csv(csv_file)
        imgnames = df['dist_img'].tolist()
        # print(imgnames)
        labels = np.array(df['type']).astype(np.float32)
        # print(labels)
        # refname = np.unique(np.array(df['ref_img']))
        # refnames_all = np.array(df['ref_img'])
        # print(refnames_all)
        sample = []
        # index = len(refname)
        for i, item in enumerate(imgnames):
            # print(refname[index[i]])
            # train_sel = (refname[index[i]] == refnames_all)
            # train_sel = np.where(train_sel == True)
            # train_sel = train_sel[0].tolist()
            # # print(train_sel)
            # for j, item in enumerate(train_sel):
            #     for aug in range(patch_num):
            sample.append((os.path.join(root, 'dist_imgs', item), labels[i]))


        self.samples = sample
        self.transform = self.img_transform(224)
        print(len(self.samples))

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])

class CSIQFolder(data.Dataset):

    def __init__(self, root, index,  patch_num, save, type_sel='all'):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []


        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                if type_sel == 'all':
                    for aug in range(patch_num):
                        sample.append((os.path.join(root, 'images', imgnames[item]), labels[item]))
                elif type_sel in imgnames[item]:
                    for aug in range(patch_num):
                        sample.append((os.path.join(root, 'images', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = self.img_transform(224)

        if save:
            header = ['image','dmos']
            csv_test = open(os.path.join(root, 'test-split.csv'), 'w')
            writer = csv.writer(csv_test)
            writer.writerow(header)

            # test_writer = csv.writer(csv_test)
            # test_writer.writerow(header)
            for i, line in enumerate(self.samples):
                writer.writerow(line)
            csv_test.close()

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])


class Koniq_10kFolder(data.Dataset):

    def __init__(self, root, index,  patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'data.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'images', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = self.img_transform(224)

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def img_transform(self, crop_size):
        normalize = Normalize(mean=[0.485, 0.456, 0.406],
                              std=[0.229, 0.224, 0.225])
        return Compose([
            Resize([512,384]),
            RandomCrop(crop_size),
            ToTensor(),
            normalize
            ])


# class BIDFolder(data.Dataset):
#
#     def __init__(self, root, index, transform, patch_num):
#
#         imgname = []
#         mos_all = []
#
#         xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
#         workbook = load_workbook(xls_file)
#         booksheet = workbook.active
#         rows = booksheet.rows
#         count = 1
#         for row in rows:
#             count += 1
#             img_num = (booksheet.cell(row=count, column=1).value)
#             img_name = "DatabaseImage%04d.JPG" % (img_num)
#             imgname.append(img_name)
#             mos = (booksheet.cell(row=count, column=2).value)
#             mos = np.array(mos)
#             mos = mos.astype(np.float32)
#             mos_all.append(mos)
#             if count == 587:
#                 break
#
#         sample = []
#         for i, item in enumerate(index):
#             for aug in range(patch_num):
#                 sample.append((os.path.join(root, imgname[item]), mos_all[item]))
#
#         self.samples = sample
#         self.transform = transform
#
#     def __getitem__(self, index):
#         """
#         Args:
#             index (int): Index
#
#         Returns:
#             tuple: (sample, target) where target is class_index of the target class.
#         """
#         path, target = self.samples[index]
#         sample = pil_loader(path)
#         sample = self.transform(sample)
#         return sample, target
#
#     def __len__(self):
#         length = len(self.samples)
#         return length
#
#
# class TID2013Folder(data.Dataset):
#
#     def __init__(self, root, index, transform, patch_num):
#         refpath = os.path.join(root, 'reference_images')
#         refname = getTIDFileName(refpath,'.bmp.BMP')
#         txtpath = os.path.join(root, 'mos_with_names.txt')
#         fh = open(txtpath, 'r')
#         imgnames = []
#         target = []
#         refnames_all = []
#         for line in fh:
#             line = line.split('\n')
#             words = line[0].split()
#             imgnames.append((words[1]))
#             target.append(words[0])
#             ref_temp = words[1].split("_")
#             refnames_all.append(ref_temp[0][1:])
#         labels = np.array(target).astype(np.float32)
#         refnames_all = np.array(refnames_all)
#
#         sample = []
#         for i, item in enumerate(index):
#             train_sel = (refname[index[i]] == refnames_all)
#             train_sel = np.where(train_sel == True)
#             train_sel = train_sel[0].tolist()
#             for j, item in enumerate(train_sel):
#                 for aug in range(patch_num):
#                     sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
#         self.samples = sample
#         self.transform = transform
#
#     def __getitem__(self, index):
#         """
#         Args:
#             index (int): Index
#
#         Returns:
#             tuple: (sample, target) where target is class_index of the target class.
#         """
#         path, target = self.samples[index]
#         sample = pil_loader(path)
#         sample = self.transform(sample)
#         return sample, target
#
#     def __len__(self):
#         length = len(self.samples)
#         return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')